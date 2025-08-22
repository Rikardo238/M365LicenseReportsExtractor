# Glaronia Informatik AG
# Rikardo Stoilov
# 22.08.2024

import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def save_as_excel(rows: list[dict], date: str, filename) -> None:
    if not rows:
        print(f"⚠️ Keine Organisationsdaten gefunden!")
        sys.exit(1)

    # DataFrame bauen
    df = pd.DataFrame(rows)

    # DataFrame sortieren alphabetisch nach Firma
    df = df.sort_values(by=["Firma"], key=lambda col: col.str.lower())

    # NaN -> leere Zellen
    df = df.fillna("")

    # Excel schreiben
    out_xlsx = f"{date} {filename}.xlsx"
    df.to_excel(out_xlsx, index=False)

    # Styling
    wb = load_workbook(out_xlsx)
    ws = wb.active
    ws.title = f"{date}"

    # Header stylen
    header_fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
    header_font = Font(bold=True)
    thin = Side(border_style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    # Zebra-Fill für jede zweite Zeile
    zebra_fill = PatternFill(start_color="FFF7F7F7", end_color="FFF7F7F7", fill_type="solid")

    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            # Zebra-Style: jede gerade Zeile färben
            if r % 2 == 0:
                cell.fill = zebra_fill
            # Zahlenzentriert, Text links
            if isinstance(cell.value, (int, float)) and str(cell.value).isdigit():
                cell.alignment = center
                cell.number_format = "#,##0"
            else:
                cell.alignment = left

    # AutoFilter & Freeze Pane
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Auto-fit Spaltenbreite
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 50)

    wb.save(out_xlsx)
    print(f"Excel gespeichert: {out_xlsx}")
